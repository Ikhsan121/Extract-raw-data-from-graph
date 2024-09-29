import pandas as pd
from openpyxl.reader.excel import load_workbook
import os

directory_path = "./Results"

try:
    # Try to create the directory
    os.makedirs(directory_path)
except FileExistsError:
    # This will be raised if the directory already exists
    print(f"Directory '{directory_path}' already exists.")
except PermissionError:
    # This will be raised if you don't have permission to create the directory
    print(f"You do not have the permissions to create '{directory_path}'.")
except Exception as e:
    # This will catch any other unexpected errors
    print(f"An unexpected error occurred: {e}")

def create_excel(filename, response):

    df1 = pd.DataFrame(response['nav_TR_API_json'])
    df2 = pd.DataFrame(response['share_price_TR_API_json'])
    # df3 and df4 for calculate discount
    df3 = pd.DataFrame(response['nav_API_json'])
    df4 = pd.DataFrame(response['price_API_json'])

    # rename column
    df1.columns = ['EndDate', 'Value_1']
    df2.columns = ['EndDate', 'Value_2']
    df3.columns = ['EndDate','OriginalDate' ,'Value_3']
    df3 = df3.drop('OriginalDate', axis=1) # delete second column
    df4.columns = ['EndDate', 'OriginalDate','Value_4']
    df4 = df4.drop('OriginalDate', axis=1)

    # Convert 'Value' column to numeric in case it's stored as strings
    df1['Value_1'] = pd.to_numeric(df1['Value_1'], errors='coerce')
    df2['Value_2'] = pd.to_numeric(df2['Value_2'], errors='coerce')
    df3['Value_3'] = pd.to_numeric(df3['Value_3'], errors='coerce')
    df4['Value_4'] = pd.to_numeric(df4['Value_4'], errors='coerce')

    # merging df1 and df2
    merged_df = pd.merge(df1, df2, on='EndDate')

    # Renaming columns to reflect their respective sources
    merged_df.columns = ['Date', 'Blue', 'Red']

    # formula for discount
    df4['Green']= (((df4['Value_4'] - df3['Value_3']) / df3['Value_3']) * 100)
    df4 = df4.drop('Value_4', axis=1)
    df4.columns = ['Date', 'Green']

    # merging df1 and df4 to add discount columns
    final_df = pd.merge(merged_df, df4, on='Date')
    # Write DataFrame to Excel starting from the second row
    final_df.to_excel(f'./Results/{filename}.xlsx', index=False, startrow=1)

    # Load the workbook and the active worksheet
    wb = load_workbook(f'./Results/{filename}.xlsx')
    ws = wb.active

    # Manually set headers and merge for the city-of-london-investment-trust
    ws['B1'] = filename

    # Set the second-row headers for Red, Blue, and Green
    ws['B2'] = 'Red'
    ws['C2'] = 'Blue'
    ws['D2'] = 'Green'

    # Save the workbook
    wb.save(f'./Results/{filename}.xlsx')
